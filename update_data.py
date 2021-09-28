# 要添加一个新单元，输入 '# %%'
# 要添加一个新的标记单元，输入 '# %% [markdown]'
# %%
from openpyxl import load_workbook
from docx import Document
import pandas as pd


# %%
df = pd.read_excel("订单数据.xlsx", header=None)


# %%
mapping = dict(zip(df[0], df[1]))


# %%
mapping


# %%
doc = Document(docx="订单信息.docx")


# %%
table = doc.tables[0]


# %%
type(table.rows)


# %%
list_data = []


# %%
len_ = len(table.rows)
for i in range(len_):
    row = table.rows[i]
    data = []
    for cell in row.cells:
        data.append(cell.text)
    list_data.append(data)


# %%
list_data


# %%
wb = load_workbook(filename="订单信息.xlsx")


# %%
ws = wb.active


# %%
excel_data = []
for row in list(range(1, 20)):
    row_data = []
    for col in list(range(1, 20)):
        row_data.append(ws.cell(row=row, column=col).value)
    excel_data.append(row_data)


# %%
df_excel = pd.DataFrame(excel_data).dropna(axis="index", how="all")


# %%
df_excel = df_excel.dropna(axis="columns", how="all")


# %%
df_excel


# %%
ws.cell(row=2, column=5).value == None


# %%
for row in list(range(1, 20)):
    row_data = []
    for col in list(range(1, 20)):
        val = ws.cell(row=row, column=col).value
        cell_right = ws.cell(row=row, column=col+1)
        crval = cell_right.value
        if crval == None:
            if val in mapping:
                ws.cell(row=row, column=col+1, value = mapping[val])
        else:
            mapping[val]=crval


# %%
mapping


# %%
wb.save("out/订单信息.xlsx")


# %%
df2=pd.DataFrame(list_data)


# %%
df2


# %%
len(df2.columns)


# %%
len(df2.index)


# %%
mapping


# %%
for i in range(len(table.rows)):
    for j in range(len(table.rows[i].cells)):
        text=table.cell(i, j).text
        if text in mapping:
            print(text)
            try:
                table.cell(i, j+1).text=str(mapping[text])
            except:
                pass


# %%
doc.save("out/订单信息.docx")


# %%
